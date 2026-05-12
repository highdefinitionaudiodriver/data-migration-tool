#!/usr/bin/env python3
"""DataMigrationTool 設計書生成スクリプト"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# --- 共通スタイル ---
HEADER_FONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
SUB_HEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
SUB_HEADER_FONT = Font(name="Arial", bold=True, size=10)
CELL_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="2F5496")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, align=None):
    cell = ws.cell(row=row, column=col)
    cell.font = CELL_FONT
    cell.border = THIN_BORDER
    cell.alignment = align or WRAP
    return cell


def auto_width(ws, col_widths):
    for col_idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


# ============================================================
# Sheet 1: 機能一覧表（基本設計）
# ============================================================
ws1 = wb.active
ws1.title = "機能一覧表"
ws1.sheet_properties.tabColor = "2F5496"

ws1.merge_cells("A1:G1")
ws1["A1"] = "DataMigrationTool - 機能一覧表（基本設計）"
ws1["A1"].font = TITLE_FONT

headers1 = ["機能ID", "機能名", "機能カテゴリ", "概要", "対象ユーザー", "入力/出力", "実装ファイル"]
for c, h in enumerate(headers1, 1):
    ws1.cell(row=3, column=c, value=h)
style_header_row(ws1, 3, len(headers1))

data1 = [
    ["F-001", "設定ファイル読み込み", "設定管理",
     "外部JSONファイル(mapping_config.json)からマッピング定義を読み込み、フィールド名・型・制約・DB型情報を検証する。",
     "全ユーザー", "入力: mapping_config.json",
     "migrate.py: load_config()\nDataMigrationTool.java: loadConfig()"],
    ["F-002", "文字コード自動判定", "入力処理",
     "入力ファイルのバイト列を解析し、UTF-8(BOM有無)・UTF-16・Shift-JIS(CP932)・EUC-JPを自動判定する。chardetライブラリがあればフォールバックとして使用。",
     "全ユーザー", "入力: 任意のテキストファイル",
     "migrate.py: detect_encoding()\nDataMigrationTool.java: detectEncoding()"],
    ["F-003", "CSV/TSV読み込み", "入力処理",
     "CSVまたはTSVファイルを読み込む。区切り文字(カンマ/タブ/セミコロン/パイプ)を先頭4096バイトから自動判定し、DictReaderで辞書リストに変換する。",
     "全ユーザー", "入力: .csv / .tsv ファイル",
     "migrate.py: read_csv()\nDataMigrationTool.java: readCsv()"],
    ["F-004", "JSON読み込み", "入力処理",
     "JSONファイルを読み込む。トップレベルが配列の場合はそのまま、オブジェクトの場合は'records'キーの配列を取得する。",
     "全ユーザー", "入力: .json ファイル",
     "migrate.py: read_json_input()\nDataMigrationTool.java: readJson()"],
    ["F-005", "XML読み込み", "入力処理",
     "XMLファイルを読み込む。ルート要素の直下の各子要素を1レコードとして扱い、その子要素のタグ名をキー、テキストを値とする辞書に変換する。Java版はXXE攻撃対策済み。",
     "全ユーザー", "入力: .xml ファイル",
     "migrate.py: read_xml_input()\nDataMigrationTool.java: readXml()"],
    ["F-006", "バリデーション & 型変換", "データ変換",
     "設定ファイルの定義に基づき、各フィールドの必須チェック・型変換(str/int/float/date/bool)・文字数上限チェックを実行する。日付は5形式を自動認識し、真偽値は日本語表記(○/はい等)にも対応。",
     "全ユーザー", "内部処理",
     "migrate.py: convert_value(), transform_record()\nDataMigrationTool.java: convertValue(), transformRecord()"],
    ["F-007", "JSON出力", "出力処理",
     "変換済みレコードを整形済みJSON(UTF-8, indent=2)としてファイルに書き出す。WebアプリAPI連携用途を想定。",
     "全ユーザー", "出力: .json ファイル",
     "migrate.py: output_json()\nDataMigrationTool.java: outputJson()"],
    ["F-008", "Oracle SQL出力", "出力処理",
     "Oracle固有構文でINSERT文を生成する。文字列はN'...'リテラル、日付はTO_DATE()関数、真偽値はNUMBER(1)の1/0で出力。末尾にCOMMITを付与。",
     "DBA / 開発者", "出力: .sql ファイル",
     "migrate.py: output_sql(dialect='oracle')\nDataMigrationTool.java: outputSql()"],
    ["F-009", "PostgreSQL SQL出力", "出力処理",
     "PostgreSQL固有構文でINSERT文を生成する。日付は'...'::DATEキャスト、真偽値はTRUE/FALSE。BEGIN/COMMITでトランザクション囲み。",
     "DBA / 開発者", "出力: .sql ファイル",
     "migrate.py: output_sql(dialect='postgresql')\nDataMigrationTool.java: outputSql()"],
    ["F-010", "CREATE TABLE文生成", "出力処理",
     "設定ファイルのoracle_type/pg_type定義に基づき、対象DB方言のCREATE TABLE DDL文を参考用として出力する。",
     "DBA / 開発者", "出力: .sql ファイル",
     "migrate.py: output_create_table()\nDataMigrationTool.java: outputCreateTable()"],
    ["F-011", "エラーログ出力", "運用支援",
     "バリデーション失敗した行をスキップし、行番号とエラー内容をerrors.logに出力する。処理全体は中断せず、最後にサマリー(全件/成功/スキップ)をログ表示する。",
     "全ユーザー", "出力: errors.log",
     "migrate.py: run()\nDataMigrationTool.java: main()"],
    ["F-012", "CLIインターフェース", "ユーザーI/F",
     "argparse(Python)/独自パーサー(Java)によるCLI引数処理。必須の設定ファイル(-c)、出力先(-o)、形式(-f)、DB方言(-d)、テーブル名(-t)、詳細ログ(-v)等のオプションを提供。",
     "全ユーザー", "CLI引数",
     "migrate.py: main()\nDataMigrationTool.java: main()"],
]

for r, row_data in enumerate(data1, 4):
    for c, val in enumerate(row_data, 1):
        cell = style_data_cell(ws1, r, c)
        cell.value = val
    ws1.cell(row=r, column=1).alignment = CENTER

auto_width(ws1, [10, 22, 14, 55, 16, 22, 40])


# ============================================================
# Sheet 2: CLI仕様書（本ツールはAPIではなくCLIツール）
# ============================================================
ws2 = wb.create_sheet("CLI仕様書")
ws2.sheet_properties.tabColor = "548235"

ws2.merge_cells("A1:H1")
ws2["A1"] = "DataMigrationTool - CLI仕様書（詳細設計）"
ws2["A1"].font = TITLE_FONT

headers2 = ["引数ID", "オプション名", "短縮形", "必須", "デフォルト値", "型", "説明", "バリデーション"]
for c, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=c, value=h)
style_header_row(ws2, 3, len(headers2))

data2 = [
    ["ARG-001", "input", "(位置引数)", "必須", "-", "文字列(パス)",
     "入力ファイルパス。CSV/TSV/JSON/XMLに対応。拡張子で形式を自動判別する。",
     "ファイル存在チェック。未対応拡張子はValueError。"],
    ["ARG-002", "--config", "-c", "必須", "-", "文字列(パス)",
     "マッピング設定ファイル(JSON)のパス。フィールド定義・テーブル名・型情報を記載。",
     "ファイル存在チェック。fieldsキー必須。各フィールドにsource必須。typeは5種のみ許可。"],
    ["ARG-003", "--output", "-o", "任意", "output.json", "文字列(パス)",
     "出力ファイルパス。", "-"],
    ["ARG-004", "--format", "-f", "任意", "json", "json | sql",
     "出力形式。jsonはJSON配列、sqlはINSERT文を生成。",
     "choices=[json, sql]。不正値はargparseがエラー。"],
    ["ARG-005", "--dialect", "-d", "任意", "postgresql", "oracle | postgresql",
     "SQL出力時のDB方言。Oracle固有構文とPostgreSQL固有構文を切り替える。",
     "choices=[oracle, postgresql]。"],
    ["ARG-006", "--error-log", "-e", "任意", "errors.log", "文字列(パス)",
     "バリデーションエラーのログ出力先ファイルパス。", "-"],
    ["ARG-007", "--table", "-t", "任意", "設定ファイルのtable_name", "文字列",
     "SQL出力時のテーブル名。指定時は設定ファイルの値を上書きする。",
     "-"],
    ["ARG-008", "--create-table", "-", "任意", "出力しない", "文字列(パス)",
     "CREATE TABLE DDL文の出力先。省略時はDDLを生成しない。", "-"],
    ["ARG-009", "--verbose", "-v", "任意", "False", "フラグ",
     "詳細ログモード。logging.DEBUGレベルの出力を有効にする。", "-"],
]

for r, row_data in enumerate(data2, 4):
    for c, val in enumerate(row_data, 1):
        cell = style_data_cell(ws2, r, c)
        cell.value = val
    ws2.cell(row=r, column=1).alignment = CENTER
    ws2.cell(row=r, column=4).alignment = CENTER

auto_width(ws2, [10, 16, 10, 8, 24, 18, 50, 42])


# ============================================================
# Sheet 3: テーブル定義書（詳細設計）
# ============================================================
ws3 = wb.create_sheet("テーブル定義書")
ws3.sheet_properties.tabColor = "BF8F00"

ws3.merge_cells("A1:I1")
ws3["A1"] = "DataMigrationTool - テーブル定義書（詳細設計）"
ws3["A1"].font = TITLE_FONT

ws3["A3"] = "テーブル定義はmapping_config.jsonにより動的に決定されます。以下はサンプル設定ファイル(m_employee)に基づく定義です。"
ws3["A3"].font = Font(name="Arial", size=10, italic=True, color="666666")
ws3.merge_cells("A3:I3")

headers3 = ["No.", "カラム名(物理)", "カラム名(論理)", "Oracle型", "PostgreSQL型",
            "NOT NULL", "デフォルト値", "バリデーション型", "最大文字数"]
for c, h in enumerate(headers3, 1):
    ws3.cell(row=5, column=c, value=h)
style_header_row(ws3, 5, len(headers3))

# テーブル名表示
ws3["A7"] = "テーブル名:"
ws3["A7"].font = SUB_HEADER_FONT
ws3["B7"] = "m_employee"
ws3["B7"].font = Font(name="Arial", size=10, bold=True)
ws3.merge_cells("A7:B7")

table_data = [
    [1, "emp_id", "社員番号", "VARCHAR2(20)", "VARCHAR(20)", "YES", "-", "str", 20],
    [2, "full_name", "氏名", "NVARCHAR2(100)", "VARCHAR(100)", "YES", "-", "str", 100],
    [3, "department", "部署", "NVARCHAR2(50)", "VARCHAR(50)", "NO", "null", "str", 50],
    [4, "email", "メールアドレス", "NVARCHAR2(254)", "VARCHAR(254)", "NO", "null", "str", 254],
    [5, "hire_date", "入社年月日", "DATE", "DATE", "NO", "null", "date", "-"],
    [6, "salary", "基本給", "NUMBER(10)", "INTEGER", "NO", "0", "int", "-"],
    [7, "rating", "評価スコア", "NUMBER(5,2)", "NUMERIC(5,2)", "NO", "null", "float", "-"],
    [8, "is_active", "在籍フラグ", "NUMBER(1)", "BOOLEAN", "NO", "true", "bool", "-"],
]

for r, row_data in enumerate(table_data, 8):
    for c, val in enumerate(row_data, 1):
        cell = style_data_cell(ws3, r, c)
        cell.value = val
    ws3.cell(row=r, column=1).alignment = CENTER
    ws3.cell(row=r, column=6).alignment = CENTER
    ws3.cell(row=r, column=9).alignment = CENTER

# 設定ファイルスキーマ説明
schema_start = len(table_data) + 10
ws3.cell(row=schema_start, column=1, value="設定ファイル(mapping_config.json)のフィールド定義スキーマ").font = SUB_HEADER_FONT
ws3.merge_cells(f"A{schema_start}:I{schema_start}")

schema_headers = ["プロパティ名", "型", "必須", "デフォルト", "説明"]
for c, h in enumerate(schema_headers, 1):
    ws3.cell(row=schema_start + 1, column=c, value=h)
style_header_row(ws3, schema_start + 1, len(schema_headers))

schema_data = [
    ["source", "string", "必須", "-", "元データ(CSV等)のカラム名/ヘッダー名"],
    ["type", "string", "任意", "str", "データ型: str / int / float / date / bool"],
    ["required", "boolean", "任意", "false", "true=NOT NULL相当。空値でバリデーションエラー"],
    ["max_len", "integer", "任意", "制限なし", "str型の最大文字数。超過でバリデーションエラー"],
    ["default", "any", "任意", "null", "値が空の場合に使用するデフォルト値"],
    ["oracle_type", "string", "任意", "VARCHAR2(255)", "Oracle用CREATE TABLE型名"],
    ["pg_type", "string", "任意", "VARCHAR(255)", "PostgreSQL用CREATE TABLE型名"],
]

for r, row_data in enumerate(schema_data, schema_start + 2):
    for c, val in enumerate(row_data, 1):
        cell = style_data_cell(ws3, r, c)
        cell.value = val
    ws3.cell(row=r, column=3).alignment = CENTER

auto_width(ws3, [6, 18, 18, 18, 18, 10, 14, 16, 14])


# ============================================================
# Sheet 4: エラー・ログ定義書（詳細設計）
# ============================================================
ws4 = wb.create_sheet("エラー・ログ定義書")
ws4.sheet_properties.tabColor = "C00000"

ws4.merge_cells("A1:G1")
ws4["A1"] = "DataMigrationTool - エラー・ログ定義書（詳細設計）"
ws4["A1"].font = TITLE_FONT

# --- ログレベル定義 ---
ws4["A3"] = "ログレベル定義"
ws4["A3"].font = SUB_HEADER_FONT

log_headers = ["レベル", "用途", "出力先", "例"]
for c, h in enumerate(log_headers, 1):
    ws4.cell(row=4, column=c, value=h)
style_header_row(ws4, 4, len(log_headers))

log_data = [
    ["DEBUG", "詳細なデバッグ情報（-v指定時のみ出力）", "標準エラー出力(stderr)", "-"],
    ["INFO", "正常系の処理状況通知", "標準エラー出力(stderr)",
     "設定ファイル読み込み完了 / 読み込み件数 / 出力完了 / サマリー"],
    ["WARNING", "処理続行可能な問題", "標準エラー出力(stderr)",
     "エラー件数: N -> errors.log"],
    ["ERROR", "処理中断を伴う致命的エラー", "標準エラー出力(stderr)",
     "入力ファイルが見つかりません / 設定ファイルエラー"],
]

for r, row_data in enumerate(log_data, 5):
    for c, val in enumerate(row_data, 1):
        cell = style_data_cell(ws4, r, c)
        cell.value = val
    ws4.cell(row=r, column=1).alignment = CENTER

# --- バリデーションエラー定義 ---
err_start = 11
ws4.cell(row=err_start, column=1, value="バリデーションエラー定義（errors.log出力）").font = SUB_HEADER_FONT

err_headers = ["エラーID", "エラー種別", "対象型", "エラーメッセージ形式", "出力タイミング", "出力先", "処理への影響"]
for c, h in enumerate(err_headers, 1):
    ws4.cell(row=err_start + 1, column=c, value=h)
style_header_row(ws4, err_start + 1, len(err_headers))

err_data = [
    ["VE-001", "必須項目欠損", "全型",
     "行N: 'フィールド名' は必須項目です",
     "required=trueのフィールドが空値またはnullの場合", "errors.log", "該当行をスキップ"],
    ["VE-002", "最大文字数超過", "str",
     "行N: 'フィールド名' が最大長 M を超えています (L文字)",
     "max_len定義を超える文字数の値が入力された場合", "errors.log", "該当行をスキップ"],
    ["VE-003", "整数変換失敗", "int",
     "行N: 'フィールド名' を整数に変換できません: '値'",
     "カンマ除去後もint()で変換不能な場合", "errors.log", "該当行をスキップ"],
    ["VE-004", "浮動小数点変換失敗", "float",
     "行N: 'フィールド名' を数値に変換できません: '値'",
     "カンマ除去後もfloat()で変換不能な場合", "errors.log", "該当行をスキップ"],
    ["VE-005", "日付形式不正", "date",
     "行N: 'フィールド名' の日付形式を認識できません: '値'",
     "5種の日付パターン全てにマッチしない場合", "errors.log", "該当行をスキップ"],
    ["VE-006", "真偽値変換失敗", "bool",
     "行N: 'フィールド名' を真偽値に変換できません: '値'",
     "true/false/1/0/yes/no/はい/いいえ等に該当しない場合", "errors.log", "該当行をスキップ"],
]

for r, row_data in enumerate(err_data, err_start + 2):
    for c, val in enumerate(row_data, 1):
        cell = style_data_cell(ws4, r, c)
        cell.value = val
    ws4.cell(row=r, column=1).alignment = CENTER

# --- 設定ファイルエラー定義 ---
cfg_start = err_start + len(err_data) + 4
ws4.cell(row=cfg_start, column=1, value="設定ファイルエラー定義（起動時に検出、即時終了）").font = SUB_HEADER_FONT

cfg_headers = ["エラーID", "エラー種別", "エラーメッセージ", "出力タイミング", "処理への影響"]
for c, h in enumerate(cfg_headers, 1):
    ws4.cell(row=cfg_start + 1, column=c, value=h)
style_header_row(ws4, cfg_start + 1, len(cfg_headers))

cfg_data = [
    ["CE-001", "設定ファイル未発見", "設定ファイルが見つかりません: <パス>",
     "load_config()実行時", "FileNotFoundError → 即時終了"],
    ["CE-002", "fieldsキー欠損", "設定ファイルに 'fields' オブジェクトが必要です",
     "load_config()実行時", "ValueError → 即時終了"],
    ["CE-003", "フィールド定義空", "'fields' にフィールド定義が1つ以上必要です",
     "load_config()実行時", "ValueError → 即時終了"],
    ["CE-004", "source未定義", "フィールド '<名前>' に 'source' が未定義です",
     "load_config()実行時", "ValueError → 即時終了"],
    ["CE-005", "無効な型指定", "フィールド '<名前>' の type '<型>' は無効です",
     "load_config()実行時", "ValueError → 即時終了"],
    ["CE-006", "入力ファイル未発見", "入力ファイルが見つかりません: <パス>",
     "main()実行時", "sys.exit(1) → 即時終了"],
    ["CE-007", "未対応ファイル形式", "未対応のファイル形式: <拡張子>",
     "load_input()実行時", "ValueError → 即時終了"],
]

for r, row_data in enumerate(cfg_data, cfg_start + 2):
    for c, val in enumerate(row_data, 1):
        cell = style_data_cell(ws4, r, c)
        cell.value = val
    ws4.cell(row=r, column=1).alignment = CENTER

auto_width(ws4, [10, 20, 14, 50, 42, 18, 20])


# ============================================================
# Sheet 5: アーキテクチャ図解（Mermaidコード集）
# ============================================================
ws5 = wb.create_sheet("アーキテクチャ図解")
ws5.sheet_properties.tabColor = "7030A0"

ws5.merge_cells("A1:B1")
ws5["A1"] = "DataMigrationTool - アーキテクチャ図解（Mermaidコード集）"
ws5["A1"].font = TITLE_FONT

diagrams = [
    ("システム構成図", """graph TD
    subgraph 入力ファイル
        CSV[".csv / .tsv"]
        JSON_IN[".json"]
        XML_IN[".xml"]
    end

    subgraph 設定ファイル
        CONFIG["mapping_config.json"]
    end

    subgraph DataMigrationTool
        ENC["文字コード自動判定<br/>detect_encoding()"]
        PARSER["入力パーサー<br/>read_csv / read_json / read_xml"]
        VALIDATOR["バリデーション & 型変換<br/>convert_value() / transform_record()"]
        FORMATTER["出力フォーマッター<br/>output_json / output_sql"]
    end

    subgraph 出力
        JSON_OUT["JSON ファイル"]
        ORA_SQL["Oracle SQL<br/>INSERT文"]
        PG_SQL["PostgreSQL SQL<br/>INSERT文"]
        DDL["CREATE TABLE<br/>DDL文"]
        ERR_LOG["errors.log"]
    end

    CSV --> ENC
    JSON_IN --> ENC
    XML_IN --> ENC
    CONFIG --> VALIDATOR
    ENC --> PARSER
    PARSER --> VALIDATOR
    VALIDATOR -->|成功| FORMATTER
    VALIDATOR -->|エラー| ERR_LOG
    FORMATTER --> JSON_OUT
    FORMATTER --> ORA_SQL
    FORMATTER --> PG_SQL
    FORMATTER --> DDL"""),

    ("処理シーケンス図", """sequenceDiagram
    participant User as ユーザー
    participant CLI as CLI (argparse)
    participant Config as load_config()
    participant Input as load_input()
    participant Enc as detect_encoding()
    participant Val as transform_record()
    participant Out as output_json/sql()
    participant ErrLog as errors.log

    User->>CLI: コマンド実行<br/>DataMigrationTool input.csv -c config.json -f sql -d postgresql
    CLI->>Config: 設定ファイル読み込み
    Config-->>CLI: field_map, table_name
    CLI->>Input: 入力ファイル読み込み
    Input->>Enc: 文字コード判定
    Enc-->>Input: encoding (utf-8, cp932, etc.)
    Input-->>CLI: rows (辞書リスト)

    loop 各行に対して
        CLI->>Val: transform_record(row, index, field_map)
        alt バリデーション成功
            Val-->>CLI: 変換済みレコード
        else バリデーション失敗
            Val-->>ErrLog: エラーメッセージ出力
        end
    end

    CLI->>Out: 変換済みレコード群を出力
    Out-->>User: SQL/JSONファイル生成完了
    CLI-->>User: サマリー表示 (全N件/成功M件/スキップK件)"""),

    ("モジュール構成図 (Python版)", """graph LR
    subgraph migrate.py
        main["main()<br/>CLIエントリポイント"]
        run["run()<br/>メイン処理ループ"]
        lc["load_config()<br/>設定ファイル読み込み"]
        de["detect_encoding()<br/>文字コード判定"]
        rc["read_csv()<br/>CSV/TSVパーサー"]
        rj["read_json_input()<br/>JSONパーサー"]
        rx["read_xml_input()<br/>XMLパーサー"]
        li["load_input()<br/>拡張子ルーター"]
        cv["convert_value()<br/>型変換+バリデーション"]
        tr["transform_record()<br/>1行変換"]
        oj["output_json()<br/>JSON出力"]
        os["output_sql()<br/>SQL出力"]
        oct["output_create_table()<br/>DDL出力"]
        fsv["_format_sql_value()<br/>SQL値フォーマット"]
    end

    main --> lc
    main --> run
    run --> li
    li --> rc
    li --> rj
    li --> rx
    rc --> de
    rj --> de
    rx --> de
    run --> tr
    tr --> cv
    run --> oj
    run --> os
    run --> oct
    os --> fsv"""),

    ("モジュール構成図 (Java版)", """graph LR
    subgraph DataMigrationTool.java
        main_j["main()<br/>CLIエントリポイント"]
        lc_j["loadConfig()<br/>設定ファイル読み込み"]
        pcm["parseConfigManually()<br/>JSON設定パーサー"]
        de_j["detectEncoding()<br/>文字コード判定"]
        rc_j["readCsv()<br/>CSV/TSVパーサー"]
        rj_j["readJson()<br/>JSONパーサー"]
        rx_j["readXml()<br/>XMLパーサー"]
        li_j["loadInput()<br/>拡張子ルーター"]
        cv_j["convertValue()<br/>型変換+バリデーション"]
        tr_j["transformRecord()<br/>1行変換"]
        oj_j["outputJson()<br/>JSON出力"]
        os_j["outputSql()<br/>SQL出力"]
        oct_j["outputCreateTable()<br/>DDL出力"]
        fsv_j["formatSqlValue()<br/>SQL値フォーマット"]
    end

    subgraph SimpleJsonParser.java
        sjp["parseArray()<br/>JSON配列パーサー"]
    end

    main_j --> lc_j
    lc_j --> pcm
    main_j --> li_j
    li_j --> rc_j
    li_j --> rj_j
    li_j --> rx_j
    rj_j --> sjp
    rc_j --> de_j
    rj_j --> de_j
    rx_j --> de_j
    main_j --> tr_j
    tr_j --> cv_j
    main_j --> oj_j
    main_j --> os_j
    main_j --> oct_j
    os_j --> fsv_j"""),

    ("データ変換フロー図", """flowchart TD
    START([入力値]) --> TRIM["trim() - 前後空白除去"]
    TRIM --> NULL_CHECK{"空値 or null?"}

    NULL_CHECK -->|Yes| REQ_CHECK{"required?"}
    REQ_CHECK -->|Yes| ERR_REQ["VE-001: 必須項目欠損"]
    REQ_CHECK -->|No| DEFAULT["デフォルト値を返却"]

    NULL_CHECK -->|No| TYPE_SWITCH{"type判定"}

    TYPE_SWITCH -->|str| LEN_CHECK{"max_len超過?"}
    LEN_CHECK -->|Yes| ERR_LEN["VE-002: 最大文字数超過"]
    LEN_CHECK -->|No| RET_STR["文字列を返却"]

    TYPE_SWITCH -->|int| INT_CONV["カンマ除去 + int()"]
    INT_CONV -->|成功| RET_INT["整数を返却"]
    INT_CONV -->|失敗| ERR_INT["VE-003: 整数変換失敗"]

    TYPE_SWITCH -->|float| FLOAT_CONV["カンマ除去 + float()"]
    FLOAT_CONV -->|成功| RET_FLOAT["浮動小数点を返却"]
    FLOAT_CONV -->|失敗| ERR_FLOAT["VE-004: 数値変換失敗"]

    TYPE_SWITCH -->|date| DATE_LOOP["5種の日付パターンで順次パース"]
    DATE_LOOP -->|マッチ| RET_DATE["YYYY-MM-DD形式で返却"]
    DATE_LOOP -->|全不一致| ERR_DATE["VE-005: 日付形式不正"]

    TYPE_SWITCH -->|bool| BOOL_MAP["true/1/yes/はい/○ = True<br/>false/0/no/いいえ/× = False"]
    BOOL_MAP -->|マッチ| RET_BOOL["真偽値を返却"]
    BOOL_MAP -->|不一致| ERR_BOOL["VE-006: 真偽値変換失敗"]"""),

    ("SQL方言比較図", """graph LR
    subgraph 共通処理
        REC["変換済みレコード"]
    end

    subgraph Oracle出力
        O_STR["文字列: N'値'"]
        O_DATE["日付: TO_DATE('値', 'YYYY-MM-DD')"]
        O_BOOL["真偽値: 1 / 0"]
        O_NUM["数値: そのまま"]
        O_NULL["NULL: NULL"]
        O_COMMIT["末尾: COMMIT;"]
    end

    subgraph PostgreSQL出力
        P_STR["文字列: '値'"]
        P_DATE["日付: '値'::DATE"]
        P_BOOL["真偽値: TRUE / FALSE"]
        P_NUM["数値: そのまま"]
        P_NULL["NULL: NULL"]
        P_TX["BEGIN; ... COMMIT;"]
    end

    REC --> O_STR
    REC --> O_DATE
    REC --> O_BOOL
    REC --> O_NUM
    REC --> O_NULL
    REC --> O_COMMIT

    REC --> P_STR
    REC --> P_DATE
    REC --> P_BOOL
    REC --> P_NUM
    REC --> P_NULL
    REC --> P_TX"""),
]

current_row = 3
for title, mermaid_code in diagrams:
    ws5.cell(row=current_row, column=1, value=title).font = SUB_HEADER_FONT
    ws5.cell(row=current_row, column=1).fill = SUB_HEADER_FILL
    ws5.merge_cells(f"A{current_row}:B{current_row}")
    current_row += 1

    ws5.cell(row=current_row, column=1, value="Mermaid記法（下記をMermaid Live Editorなどに貼り付けて図を生成してください）")
    ws5.cell(row=current_row, column=1).font = Font(name="Arial", size=9, italic=True, color="666666")
    ws5.merge_cells(f"A{current_row}:B{current_row}")
    current_row += 1

    code_cell = ws5.cell(row=current_row, column=1, value=mermaid_code)
    code_cell.font = Font(name="Consolas", size=9)
    code_cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws5.merge_cells(f"A{current_row}:B{current_row}")
    line_count = mermaid_code.count("\n") + 1
    ws5.row_dimensions[current_row].height = max(15 * line_count, 100)
    current_row += 2

auto_width(ws5, [80, 80])


# --- 保存 ---
output_path = "G:/マイドライブ/claudecode/data-migration-tool/design_document.xlsx"
wb.save(output_path)
print(f"設計書を生成しました: {output_path}")
